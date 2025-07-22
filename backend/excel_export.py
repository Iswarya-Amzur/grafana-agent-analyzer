#!/usr/bin/env python3
"""
Excel Export Module for LLM Analysis Reports

This module converts LLM-generated markdown reports to structured Excel files with:
- Infrastructure Widgets sheet
- System Widgets sheet  
- Summary sheet
- Formatted tables with colors and charts

Usage:
    from excel_export import ExcelReportGenerator
    
    generator = ExcelReportGenerator()
    excel_file = generator.generate_excel_report(analysis_result, widgets_data)
"""

import pandas as pd
import re
from datetime import datetime
from typing import Dict, List, Any, Tuple
import os
from dataclasses import dataclass
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Install with: pip install openpyxl")

from llm_analysis import WidgetData


@dataclass
class WidgetTableRow:
    """Structured widget data for Excel export"""
    widget_name: str
    category: str
    current_value: str
    min_value: str
    max_value: str
    trend: str
    status: str
    key_observations: str


class MarkdownTableParser:
    """Parser for extracting tables from markdown reports"""
    
    @staticmethod
    def parse_widget_table(markdown_text: str) -> List[WidgetTableRow]:
        """
        Parse the Per-Widget Analysis Table from markdown
        """
        widget_rows = []
        
        try:
            lines = markdown_text.split('\n')
            in_table = False
            header_found = False
            
            for line in lines:
                line = line.strip()
                
                # Find table start
                if "Per-Widget Analysis Table" in line:
                    in_table = True
                    continue
                    
                # Skip header and separator rows
                if in_table and "|" in line:
                    if "Widget Name" in line:
                        header_found = True
                        continue
                    if header_found and line.startswith("|---"):
                        continue
                    if header_found and line.count("|") >= 7:
                        # Parse table row
                        parts = [p.strip() for p in line.split("|")[1:-1]]  # Remove empty first/last
                        if len(parts) >= 7:
                            # Split min/max values
                            min_max = parts[3]
                            if "/" in min_max:
                                min_val, max_val = min_max.split("/", 1)
                            else:
                                min_val, max_val = "", min_max
                            
                            widget_row = WidgetTableRow(
                                widget_name=parts[0].strip(),
                                category=parts[1].strip(),
                                current_value=parts[2].strip(),
                                min_value=min_val.strip(),
                                max_value=max_val.strip(),
                                trend=parts[4].strip(),
                                status=parts[5].strip(),
                                key_observations=parts[6].strip()
                            )
                            widget_rows.append(widget_row)
                
                # End of table
                elif in_table and line.startswith("##"):
                    break
                    
        except Exception as e:
            print(f"Error parsing widget table: {e}")
            
        return widget_rows
    
    @staticmethod
    def parse_summary_metrics(markdown_text: str) -> Dict[str, Any]:
        """
        Extract key metrics from the summary section
        """
        metrics = {}
        
        try:
            lines = markdown_text.split('\n')
            
            for line in lines:
                line = line.strip()
                
                # Extract key metrics
                if "Total Widgets Analyzed" in line:
                    match = re.search(r'(\d+)', line)
                    if match:
                        metrics['total_widgets'] = int(match.group(1))
                
                elif "Widgets with Issues" in line:
                    match = re.search(r'(\d+)', line)
                    if match:
                        metrics['widgets_with_issues'] = int(match.group(1))
                
                elif "Critical Alerts" in line:
                    match = re.search(r'(\d+)', line)
                    if match:
                        metrics['critical_alerts'] = int(match.group(1))
                
                elif "Performance Score" in line:
                    match = re.search(r'(\d+(?:\.\d+)?)/10', line)
                    if match:
                        metrics['performance_score'] = float(match.group(1))
                
                elif "Model Used" in line:
                    match = re.search(r'Model Used:\s*(\S+)', line)
                    if match:
                        metrics['model_used'] = match.group(1)
                
                elif "Tokens Consumed" in line:
                    match = re.search(r'(\d+)', line)
                    if match:
                        metrics['tokens_consumed'] = int(match.group(1))
                        
        except Exception as e:
            print(f"Error parsing summary metrics: {e}")
            
        return metrics


class ExcelReportGenerator:
    """
    Generate structured Excel reports from LLM analysis
    """
    
    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_excel_report(self, analysis_result: Dict[str, Any], 
                            widgets_data: List[WidgetData]) -> str:
        """
        Generate Excel report with Infrastructure, System, and Summary sheets
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        try:
            # Parse markdown to extract structured data
            markdown_text = analysis_result.get("analysis", {}).get("raw_markdown", "")
            widget_rows = MarkdownTableParser.parse_widget_table(markdown_text)
            summary_metrics = MarkdownTableParser.parse_summary_metrics(markdown_text)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y-%m-%d")
            filename = f"grafana_report_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Create Infrastructure sheet
            self._create_infrastructure_sheet(wb, widget_rows)
            
            # Create System sheet
            self._create_system_sheet(wb, widget_rows)
            
            # Create Summary sheet (and set as active)
            self._create_summary_sheet(wb, summary_metrics, widget_rows, analysis_result)
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Set Summary as active sheet
            wb.active = wb["Summary"]
            
            # Save file
            wb.save(filepath)
            
            print(f"✅ Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            raise Exception(f"Failed to generate Excel report: {str(e)}")
    
    def _create_infrastructure_sheet(self, wb: Workbook, widget_rows: List[WidgetTableRow]):
        """Create Infrastructure Widgets sheet"""
        ws = wb.create_sheet("Infrastructure Widgets")
        
        # Filter infrastructure widgets
        infra_widgets = [w for w in widget_rows if w.category.lower() in 
                        ['infrastructure', 'storage', 'network', 'database']]
        
        if not infra_widgets:
            ws.cell(row=1, column=1, value="No infrastructure widgets found")
            return
        
        # Convert to DataFrame
        df = pd.DataFrame([{
            'Widget Name': w.widget_name,
            'Category': w.category,
            'Current Value': w.current_value,
            'Min Value': w.min_value,
            'Max Value': w.max_value,
            'Trend': w.trend,
            'Status': w.status,
            'Key Observations': w.key_observations
        } for w in infra_widgets])
        
        # Write DataFrame to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._format_widget_sheet(ws, len(infra_widgets))
    
    def _create_system_sheet(self, wb: Workbook, widget_rows: List[WidgetTableRow]):
        """Create System Widgets sheet"""
        ws = wb.create_sheet("System Widgets")
        
        # Filter system widgets
        system_widgets = [w for w in widget_rows if w.category.lower() in 
                         ['system', 'performance', 'application', 'container', 'api', 
                          'observability', 'reliability', 'security']]
        
        if not system_widgets:
            ws.cell(row=1, column=1, value="No system widgets found")
            return
        
        # Convert to DataFrame
        df = pd.DataFrame([{
            'Widget Name': w.widget_name,
            'Category': w.category,
            'Current Value': w.current_value,
            'Min Value': w.min_value,
            'Max Value': w.max_value,
            'Trend': w.trend,
            'Status': w.status,
            'Key Observations': w.key_observations
        } for w in system_widgets])
        
        # Write DataFrame to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._format_widget_sheet(ws, len(system_widgets))
    
    def _create_summary_sheet(self, wb: Workbook, summary_metrics: Dict[str, Any], 
                            widget_rows: List[WidgetTableRow], analysis_result: Dict[str, Any]):
        """Create Summary sheet with key metrics and charts"""
        ws = wb.create_sheet("Summary", 0)  # Insert at beginning
        
        # Title
        ws.cell(row=1, column=1, value="📊 Grafana Monitoring Analysis Summary").font = Font(size=16, bold=True)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Key Metrics Section
        ws.cell(row=4, column=1, value="🎯 Key Metrics").font = Font(size=14, bold=True)
        
        metrics_data = [
            ["Total Widgets Analyzed", summary_metrics.get('total_widgets', len(widget_rows))],
            ["Widgets with Issues", summary_metrics.get('widgets_with_issues', 0)],
            ["Critical Alerts", summary_metrics.get('critical_alerts', 0)],
            ["Performance Score", f"{summary_metrics.get('performance_score', 0)}/10"],
            ["Model Used", summary_metrics.get('model_used', 'GPT-4o')],
            ["Tokens Consumed", summary_metrics.get('tokens_consumed', 0)]
        ]
        
        for i, (metric, value) in enumerate(metrics_data, start=5):
            ws.cell(row=i, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Status Distribution
        ws.cell(row=12, column=1, value="⚠️ Status Distribution").font = Font(size=14, bold=True)
        
        status_counts = {}
        for widget in widget_rows:
            status = widget.status
            status_counts[status] = status_counts.get(status, 0) + 1
        
        row = 13
        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=status).font = Font(bold=True)
            ws.cell(row=row, column=2, value=count)
            
            # Color code based on status
            if status.lower() == 'critical':
                ws.cell(row=row, column=1).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif status.lower() == 'warning':
                ws.cell(row=row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif status.lower() == 'info':
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
            
            row += 1
        
        # Category Distribution
        ws.cell(row=row + 1, column=1, value="📊 Category Distribution").font = Font(size=14, bold=True)
        
        category_counts = {}
        for widget in widget_rows:
            category = widget.category
            category_counts[category] = category_counts.get(category, 0) + 1
        
        row += 2
        for category, count in category_counts.items():
            ws.cell(row=row, column=1, value=category).font = Font(bold=True)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Recommendations section
        recommendations = analysis_result.get("analysis", {}).get("structured_data", {}).get("recommendations", {})
        if recommendations:
            ws.cell(row=row + 2, column=1, value="🛠️ Recommendations Summary").font = Font(size=14, bold=True)
            row += 3
            
            for rec_type, rec_list in recommendations.items():
                if rec_list:
                    ws.cell(row=row, column=1, value=f"{rec_type.replace('_', ' ').title()}:").font = Font(bold=True)
                    row += 1
                    for rec in rec_list[:3]:  # Show top 3 recommendations
                        ws.cell(row=row, column=1, value=f"• {rec}")
                        row += 1
                    row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_widget_sheet(self, ws, row_count: int):
        """Apply formatting to widget sheets"""
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=row_count + 1, 
                               min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
        
        # Status column conditional formatting
        status_col = 7  # Status column
        for row in range(2, row_count + 2):
            cell = ws.cell(row=row, column=status_col)
            if cell.value:
                if cell.value.lower() == 'critical':
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif cell.value.lower() == 'warning':
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif cell.value.lower() == 'info':
                    cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width


def test_excel_export():
    """Test function for Excel export"""
    try:
        # Load sample analysis result
        sample_file = "./demo_reports/analysis_metadata_20250722_162828.json"
        
        if os.path.exists(sample_file):
            with open(sample_file, 'r') as f:
                sample_data = json.load(f)
            
            # Extract analysis result
            analysis_result = sample_data.get("analysis_result", {})
            
            # Create sample widget data
            widgets_data = []
            
            # Generate Excel report
            generator = ExcelReportGenerator()
            excel_file = generator.generate_excel_report(analysis_result, widgets_data)
            
            print(f"✅ Test Excel report generated: {excel_file}")
            return True
        else:
            print("❌ Sample data file not found")
            return False
            
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False


if __name__ == "__main__":
    test_excel_export()
